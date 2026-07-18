"""تصدير جدول الكميات إلى Excel — بصيغة قابلة للنقل إلى منصة اعتماد."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .config import BRAND

HEADER_FILL = PatternFill("solid", fgColor=BRAND["primary"])
ACCENT_FILL = PatternFill("solid", fgColor=BRAND["accent"])
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_boq_xlsx(proposal: dict, path: str):
    data = proposal["data"]
    wb = Workbook()
    ws = wb.active
    ws.title = "جدول الكميات"
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:G1")
    top = ws["A1"]
    top.value = f"{BRAND['name_ar']} — جدول الكميات والأسعار | {proposal['title']} | {proposal['ref_no']}"
    top.font = Font(bold=True, size=13, color="FFFFFF")
    top.fill = HEADER_FILL
    top.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["م", "الكود", "وصف البند", "الوحدة", "الكمية", "سعر الوحدة (ر.س)", "الإجمالي (ر.س)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    row = 3
    for i, line in enumerate(data.get("boq", []), 1):
        values = [i, line.get("code", ""), line["name"], line["unit"],
                  line["qty"], line["unit_price"], line["total"]]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if col != 3 else "right",
                                       vertical="center", wrap_text=True)
            if col in (6, 7):
                cell.number_format = "#,##0.00"
        row += 1

    fin = data.get("financial", {})
    summary = [
        ("التكلفة المباشرة", fin.get("direct_cost", 0)),
        (f"المصاريف الإدارية والعمومية ({fin.get('overhead_pct', 0):g}%)", fin.get("overhead", 0)),
        (f"احتياطي المخاطر ({fin.get('risk_pct', 0):g}%)", fin.get("risk", 0)),
        (f"هامش الربح ({fin.get('profit_pct', 0):g}%)", fin.get("profit", 0)),
        ("الإجمالي قبل الضريبة", fin.get("subtotal", 0)),
        (f"ضريبة القيمة المضافة ({fin.get('vat_rate', 15):g}%)", fin.get("vat", 0)),
        ("الإجمالي النهائي شامل الضريبة", fin.get("grand_total", 0)),
    ]
    row += 1
    for label, value in summary:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell = ws.cell(row=row, column=7, value=value)
        value_cell.number_format = "#,##0.00"
        value_cell.font = Font(bold=True, size=11)
        value_cell.border = BORDER
        if "النهائي" in label:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = ACCENT_FILL
        row += 1

    widths = [6, 10, 55, 10, 10, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(path)
    return path
