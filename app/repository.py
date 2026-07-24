"""المستودع المعرفي — تحليل ملفات العروض القديمة (لعزوم أو لشركات أخرى)
واستخراج بنودها المسعّرة كأسعار سوق استرشادية، وتخزين نصوصها لتغذية التوليد.
"""
import re
from io import BytesIO
from pathlib import Path

from .file_extract import extract_text

# كلمات رؤوس الأعمدة الشائعة في جداول الكميات — مرتبة بالأولوية (الوصف قبل رقم البند)
_HDR_NAME = ("وصف", "بيان", "description", "الاعمال", "الأعمال", "البند", "item")
_HDR_UNIT = ("وحدة", "الوحدة", "unit")
_HDR_QTY = ("كمية", "الكمية", "qty", "quantity")
_HDR_RATE = ("سعر", "السعر", "افرادى", "إفرادى", "rate", "price", "فئة")


def _num(v):
    try:
        x = float(str(v).replace(",", "").strip())
        return x if x == x else None  # استبعاد NaN
    except (ValueError, TypeError):
        return None


def _clean(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _hdr_norm(s) -> str:
    """تطبيع نص الرأس: إزالة المدّات (ـ) والمسافات لمطابقة «الـــوصــــف» مع «وصف»."""
    return re.sub(r"[ـ\s]+", "", str(s or "")).lower()


def _match_col(header_cells: list, keywords: tuple) -> list[int]:
    """الأعمدة المطابقة مرتبة بأولوية الكلمات المفتاحية ثم بموقع العمود."""
    out = []
    for prio, k in enumerate(keywords):
        for i, c in enumerate(header_cells):
            text = _hdr_norm(c)
            if text and k in text and i not in out:
                out.append(i)
    return out


def parse_boq_items(content: bytes, filename: str) -> list[dict]:
    """استخراج البنود المسعّرة من ملف Excel بأي تخطيط أعمدة (اكتشاف تلقائي للرؤوس)."""
    if Path(filename).suffix.lower() not in (".xlsx", ".xlsm"):
        return []
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return []

    items = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        cols = _detect_columns(rows)
        if not cols:
            continue
        name_c, unit_c, qty_c, rate_c, start = cols
        for row in rows[start:]:
            if len(row) <= max(name_c, rate_c):
                continue
            name = _clean(row[name_c])
            rate = _num(row[rate_c])
            if not name or len(name) < 6 or not rate or rate <= 0:
                continue
            # استبعاد صفوف الإجماليات
            if any(k in name for k in ("إجمالي", "الاجمالي", "total", "Total", "TOTAL", "مجموع")):
                continue
            items.append({
                "name": name[:200],
                "unit": _clean(row[unit_c])[:20] if unit_c is not None and len(row) > unit_c else "",
                "qty": _num(row[qty_c]) if qty_c is not None and len(row) > qty_c else None,
                "unit_price": rate,
            })
    return items


def _detect_columns(rows) -> tuple | None:
    """البحث عن صف الرؤوس في أول 15 صفاً وتحديد مواقع الأعمدة."""
    for idx, row in enumerate(rows[:15]):
        cells = ["" if c is None else str(c) for c in row]
        names = _match_col(cells, _HDR_NAME)
        rates = _match_col(cells, _HDR_RATE)
        if not names or not rates:
            continue
        units = _match_col(cells, _HDR_UNIT)
        qtys = _match_col(cells, _HDR_QTY)
        # استبعاد أعمدة "الإجمالي" من مرشحي السعر (سعر الوحدة لا الإجمالي)
        rate_c = None
        for r in rates:
            h = _hdr_norm(cells[r])
            if "جمال" not in h and "amount" not in h and "total" not in h:
                rate_c = r
                break
        if rate_c is None:
            continue
        # أول عمود وصف بالأولوية، مع تفضيل العربي بين أعمدة الوصف فقط (لا أعمدة رقم البند)
        name_c = names[0]
        desc_cols = [n for n in names if any(k in _hdr_norm(cells[n]) for k in ("وصف", "بيان", "description"))]
        for n in desc_cols:
            if any("؀" <= ch <= "ۿ" for ch in cells[n]):
                name_c = n
                break
        else:
            if desc_cols:
                name_c = desc_cols[0]
        return (name_c, units[0] if units else None, qtys[0] if qtys else None, rate_c, idx + 1)
    return None


def find_relevant_repo_texts(query: str, top_n: int = 2) -> list[dict]:
    """أقرب ملفات المستودع لنص المشروع — تُضاف كسياق معرفي عند التوليد."""
    from .database import get_repo_texts
    from .similarity import _tokens
    q = _tokens(query)
    if not q:
        return []
    scored = []
    for f in get_repo_texts(limit_chars=200_000):
        overlap = len(q & _tokens(f["extracted_text"] + " " + f["filename"]))
        if overlap >= 3:
            # نمرر مقتطفاً فقط للسياق حتى لا نضخم طلب التوليد
            scored.append((overlap, {**f, "extracted_text": f["extracted_text"][:6000]}))
    scored.sort(key=lambda x: -x[0])
    return [f for _, f in scored[:top_n]]


def ingest_file(filename: str, content: bytes, source_type: str, company: str, notes: str) -> dict:
    """تحليل ملف مرفوع وتخزينه في المستودع: نص كامل + بنود مسعّرة."""
    from .database import create_repo_file
    text = extract_text(filename, content)
    items = parse_boq_items(content, filename)
    meta = {"filename": filename, "source_type": source_type, "company": company, "notes": notes}
    record = create_repo_file(meta, text, items)
    return record
